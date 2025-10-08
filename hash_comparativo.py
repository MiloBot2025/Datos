#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
hash_comparativo.py — copia exacta (DB), Stock V y reportes con estado persistente

Cambios solicitados:
- Tevelam / Disco: en Hoja 1 se mantienen también los “Menor a 5” (incluye cualquier número > 0). Se descarta 0/sin stock.
- IMSA: Hoja 1 solo con stock (texto positivo o número > 0).
- ARS_Tech: Hoja 1 se mantiene > 5.
- Diffs de precios con mapeo fijo:
    * IMSA  : fila 8.., col A=código, col G=precio
    * Tevelam: fila 11.., col A=código, col L=precio
    * Disco  : fila 8.., col A=código, col M=precio
"""

from __future__ import annotations

import csv
import json
import time
import hashlib
import re
from pathlib import Path
from datetime import datetime
from typing import Optional, Dict, Any, List, Tuple
import os

from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook, Workbook

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

# ========= CONFIG =========
BASE_DIR = Path(os.getenv("WORKDIR", "."))

# Descargas efímeras
RUTA_DESCARGA = BASE_DIR / "downloads"

# ESTADO PERSISTENTE
STATE_DIR       = BASE_DIR / "_db"
STATE_HASH_DIR  = STATE_DIR / "hash"
STATE_SNAP_DIR  = STATE_DIR / "snapshots"

# Publicación (GitHub Pages)
PUBLIC_DB_DIR      = BASE_DIR / "public_db"       # copia exacta
PUBLIC_LISTAS_DIR  = BASE_DIR / "public_listas"   # Stock V
PUBLIC_REPORTS_DIR = BASE_DIR / "public_reports"  # reportes de difs

# (opcionales efímeras locales)
REPORTS_DIR = BASE_DIR / "_reports"
SNAP_DIR    = BASE_DIR / "_snapshots"

for d in (
    RUTA_DESCARGA, STATE_DIR, STATE_HASH_DIR, STATE_SNAP_DIR,
    PUBLIC_DB_DIR, PUBLIC_LISTAS_DIR, PUBLIC_REPORTS_DIR,
    REPORTS_DIR, SNAP_DIR
):
    d.mkdir(parents=True, exist_ok=True)

# Si el hash es igual → borrar descarga
BORRAR_DUPLICADO = os.getenv("BORRAR_DUPLICADO", "true").lower() == "true"

# URLs fuentes
URL_TEVELAM = "https://drive.google.com/uc?export=download&id=1hPH3VwQDtMgx_AkC5hFCUbM2MEiwBEpT"
URL_DISCO_PRO = "https://drive.google.com/uc?id=1-aQ842Dq3T1doA-Enb34iNNzenLGkVkr&export=download"
URL_PROVEEDOR_EXTRA = "https://docs.google.com/uc?id=1JnUnrpZUniTXUafkAxCInPG7O39yrld5&export=download"

# IMSA (iframe con password)
IMSA_URL = "https://listaimsa.com.ar/lista-de-precios/"
IMSA_PASSWORD = os.getenv("IMSA_PASSWORD", "lista2021")

REQ_HEADERS = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
TIMEOUT = 60

# ========= LOG / UTILS =========
def log(msg: str) -> None:
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}")

def ts() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

# ========= HASH =========
def file_sha256(path: Path, chunk_size: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open('rb') as f:
        for chunk in iter(lambda: f.read(chunk_size), b''):
            h.update(chunk)
    return h.hexdigest()

def _hash_path(source_key: str) -> Path:
    return STATE_HASH_DIR / f"{source_key}.sha256"

def read_prev_hash(source_key: str) -> Optional[str]:
    p = _hash_path(source_key)
    if p.exists():
        try:
            return p.read_text(encoding='utf-8').strip()
        except Exception:
            return None
    return None

def write_hash(source_key: str, hexhash: str) -> None:
    _hash_path(source_key).write_text(hexhash, encoding='utf-8')

def decide_should_process(source_key: str, path: Optional[Path]) -> Tuple[bool, Optional[str]]:
    if not path or not path.exists():
        log(f"⏭️ {source_key}: no hay archivo para comparar.")
        return False, None
    try:
        new_hash = file_sha256(path)
        prev_hash = read_prev_hash(source_key)
        log(f"📇 {source_key}: nuevo={new_hash[:12]}… | previo={(prev_hash[:12] + '…') if prev_hash else 'N/A'}")
        if prev_hash == new_hash:
            log(f"⏭️ {source_key}: sin cambios (hash igual).")
            if BORRAR_DUPLICADO:
                try:
                    path.unlink(missing_ok=True)
                    log(f"🗑️ {source_key}: duplicado eliminado: {path.name}")
                except Exception as e:
                    log(f"⚠️ {source_key}: no se pudo borrar duplicado: {e}")
            return False, new_hash
        write_hash(source_key, new_hash)
        log(f"🔄 {source_key}: cambios detectados (hash distinto) → conservo y proceso.")
        return True, new_hash
    except Exception as e:
        log(f"⚠️ {source_key}: error comparando hash: {e} → proceso por las dudas.")
        return True, None

# ========= DESCARGAS =========
def download_simple(url: str, base_name: str) -> Optional[Path]:
    try:
        dst = RUTA_DESCARGA / f"{base_name}_{ts()}.xlsx"
        r = requests.get(url, headers=REQ_HEADERS, timeout=TIMEOUT)
        r.raise_for_status()
        dst.write_bytes(r.content)
        log(f"✅ Descargado: {dst.name}")
        return dst
    except Exception as e:
        log(f"❌ Error descarga {base_name}: {e}")
        return None

# ========= NORMALIZACIÓN =========
def _norm_text(s: Any) -> str:
    return (str(s) if s is not None else "").strip()

def _norm_text_lc(s: Any) -> str:
    return _norm_text(s).lower().replace("ó","o").replace("í","i").replace("á","a").replace("é","e").replace("ú","u")

def try_float(v):
    if v is None: return None
    s = _norm_text(v).replace(".", "").replace(",", ".") if isinstance(v, str) else v
    try: return float(s)
    except Exception: return None

# candidatos de encabezados
CAND_COD   = {"codigo","código","cod","id","articulo","artículo","sku","modelo"}
CAND_STOCK = {"stock","estado","disponibilidad","disponible"}
CAND_PREC  = {"precio","p. lista","p lista","plista","lista","price","valor"}
CAND_MON   = {"moneda","currency","divisa"}

def detectar_columnas(ws, max_scan_rows: int = 60):
    for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_rows, values_only=True), start=1):
        tmp = {"codigo": None, "stock": None, "precio": None, "moneda": None}
        for c_i, cell in enumerate(row, start=1):
            h = _norm_text_lc(cell)
            if not h: continue
            if not tmp["codigo"] and h in CAND_COD:   tmp["codigo"] = c_i
            elif not tmp["precio"] and h in CAND_PREC: tmp["precio"] = c_i
            elif not tmp["moneda"] and h in CAND_MON:  tmp["moneda"] = c_i
            elif not tmp["stock"] and h in CAND_STOCK: tmp["stock"] = c_i
        if tmp["codigo"]:
            return r_i, tmp
    return None, {"codigo": None, "stock": None, "precio": None, "moneda": None}

# ========= Reglas de filtrado (Hoja 1) =========
_MAYOR5_RX = re.compile(r"(mayor\s*(a|que)\s*5|>\s*5)", re.I)
_MENOR5_RX = re.compile(r"(menor\s*(a|que)\s*5|<\s*5)", re.I)
_POS_IMSA = ("con stock","en stock","disponible","hay stock","sí","si","hay")

def incluir_teve_disco(stock_raw: Any, stock_num: Any) -> bool:
    """
    Tevelam/Disco: mantener Mayor a 5 y también Menor a 5.
    Regla: incluir si texto dice 'mayor a 5' o 'menor a 5' o si número > 0.
    Excluir 0 / sin stock.
    """
    t = _norm_text_lc(stock_raw)
    if t in {"sin stock","sinstock","sin-stock","no","0","agotado","sin"}:
        return False
    n = try_float(stock_num if stock_num is not None else stock_raw)
    if n is not None:
        return n > 0
    if _MAYOR5_RX.search(t) or _MENOR5_RX.search(t):
        return True
    return False

def incluir_ars_gt5(stock_raw: Any, stock_num: Any) -> bool:
    """ARS_Tech: incluir solo si > 5 (número ≥ 5 o texto 'mayor a 5')."""
    t = _norm_text_lc(stock_raw)
    n = try_float(stock_num if stock_num is not None else stock_raw)
    if n is not None:
        return n >= 5
    return bool(_MAYOR5_RX.search(t))

def incluir_imsa(stock_raw: Any, stock_num: Any) -> bool:
    """IMSA: solo con stock (texto positivo o número > 0)."""
    t = _norm_text_lc(stock_raw)
    if any(tok in t for tok in _POS_IMSA):
        return True
    n = try_float(stock_num if stock_num is not None else stock_raw)
    return (n is not None) and (n > 0)

# ========= EXTRACTORES PARA HOJA 1 (Stock V) =========
def extraer_registros_con_stock_fallback(path: Path, fila_inicio: int, col_stock: int) -> List[Dict[str, Any]]:
    """Devuelve registros con columnas: ID, Stock (clasif 0/2/6), Precio, Moneda, StockRaw, StockNum"""
    out: List[Dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row, cols = detectar_columnas(ws)
        if header_row and cols["codigo"]:
            max_needed_col = max(v for v in cols.values() if v)
            for row in ws.iter_rows(min_row=header_row+1, min_col=1, max_col=max_needed_col, values_only=True):
                cod = row[cols["codigo"]-1] if cols["codigo"] else None
                if not _norm_text(cod): continue
                stock_raw = row[cols["stock"]-1] if cols["stock"] else None
                precio = row[cols["precio"]-1] if cols["precio"] else None
                moneda = row[cols["moneda"]-1] if cols["moneda"] else None
                out.append({
                    "ID": _norm_text(cod),
                    "Stock": convertir_stock_generico(stock_raw),
                    "Precio": try_float(precio),
                    "Moneda": _norm_text(moneda) or None,
                    "StockRaw": stock_raw,
                    "StockNum": try_float(stock_raw)
                })
        else:
            max_row = ws.max_row or 1
            for r in range(fila_inicio, max_row + 1):
                cod = ws.cell(row=r, column=1).value
                if not _norm_text(cod): continue
                stock_raw = ws.cell(row=r, column=col_stock).value
                out.append({
                    "ID": _norm_text(cod),
                    "Stock": convertir_stock_generico(stock_raw),
                    "Precio": None, "Moneda": None,
                    "StockRaw": stock_raw, "StockNum": try_float(stock_raw)
                })
    finally:
        wb.close()
    return out

def convertir_stock_generico(valor):
    t = _norm_text_lc(valor)
    if t in {"sin stock","sinstock","sin-stock","no","0","agotado","sin"}:
        return 0
    if t in {"menor a 5","<5","bajo","consultar","poco","limitado"}:
        return 2
    if t in {"mayor a 5",">5","alto","con stock","constock","en stock","stock","disponible","si","sí"}:
        return 6
    try:
        n = float(t.replace(",", "."))
        if n <= 0: return 0
        return 6 if n >= 5 else 2
    except Exception:
        return None

# TEVELAM Hoja1 (fila 11, stock col 9) + espejo F/T — incluir también “Menor a 5”
def extraer_tevelam_hoja1(path: Path) -> List[Dict[str, Any]]:
    regs = extraer_registros_con_stock_fallback(path, fila_inicio=11, col_stock=9)
    out = []
    for r in regs:
        if not incluir_teve_disco(r.get("StockRaw"), r.get("StockNum")):
            continue
        out.append(r)
        s = r["ID"]
        if len(s) >= 2:
            if s.startswith("F"):
                out.append({"ID": "T"+s[1:], "Stock": r["Stock"], "Precio": None, "Moneda": None,
                            "StockRaw": r.get("StockRaw"), "StockNum": r.get("StockNum")})
            elif s.startswith("T"):
                out.append({"ID": "F"+s[1:], "Stock": r["Stock"], "Precio": None, "Moneda": None,
                            "StockRaw": r.get("StockRaw"), "StockNum": r.get("StockNum")})
    return out

# DISCO PRO Hoja1 (fila 8, stock col 7) + espejo — incluir también “Menor a 5”
def extraer_disco_hoja1(path: Path) -> List[Dict[str, Any]]:
    regs = extraer_registros_con_stock_fallback(path, fila_inicio=8, col_stock=7)
    out = []
    for r in regs:
        if not incluir_teve_disco(r.get("StockRaw"), r.get("StockNum")):
            continue
        out.append(r)
        s = r["ID"]
        if len(s) >= 2:
            if s.startswith("F"):
                out.append({"ID": "T"+s[1:], "Stock": r["Stock"], "Precio": None, "Moneda": None,
                            "StockRaw": r.get("StockRaw"), "StockNum": r.get("StockNum")})
            elif s.startswith("T"):
                out.append({"ID": "F"+s[1:], "Stock": r["Stock"], "Precio": None, "Moneda": None,
                            "StockRaw": r.get("StockRaw"), "StockNum": r.get("StockNum")})
    return out

# PROVEEDOR EXTRA Hoja1 — > 5 (sin cambios)
def extraer_extra_hoja1(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out: List[Dict[str, Any]] = []
        if "STOCK" in wb.sheetnames:
            st = wb["STOCK"]
            header_row, cols = detectar_columnas(st)
            if header_row:
                max_needed_col = max(v for v in cols.values() if v)
                for row in st.iter_rows(min_row=header_row+1, min_col=1, max_col=max_needed_col, values_only=True):
                    cod = row[cols["codigo"]-1] if cols["codigo"] else None
                    if not _norm_text(cod): continue
                    stock_raw = row[cols["stock"]-1] if cols["stock"] else None
                    precio = row[cols["precio"]-1] if cols["precio"] else None
                    moneda = row[cols["moneda"]-1] if cols["moneda"] else None
                    if incluir_ars_gt5(stock_raw, try_float(stock_raw)):
                        out.append({
                            "ID": _norm_text(cod),
                            "Stock": convertir_stock_generico(stock_raw),
                            "Precio": try_float(precio),
                            "Moneda": _norm_text(moneda) or None,
                            "StockRaw": stock_raw, "StockNum": try_float(stock_raw)
                        })
            else:
                max_row = st.max_row or 1
                for r in range(2, max_row + 1):
                    _id = st.cell(row=r, column=2).value
                    if not _norm_text(_id): continue
                    stock_raw = st.cell(row=r, column=4).value
                    if incluir_ars_gt5(stock_raw, try_float(stock_raw)):
                        out.append({"ID": _norm_text(_id),
                                    "Stock": convertir_stock_generico(stock_raw),
                                    "Precio": None, "Moneda": None,
                                    "StockRaw": stock_raw, "StockNum": try_float(stock_raw)})
            return out
        else:
            regs = extraer_registros_con_stock_fallback(path, fila_inicio=2, col_stock=8)
            return [r for r in regs if incluir_ars_gt5(r.get("StockRaw"), r.get("StockNum"))]
    finally:
        wb.close()

# IMSA Hoja1 — solo con stock
def extraer_imsa_hoja1(path: Path) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        header_row, cols = detectar_columnas(ws)
        if header_row and cols["codigo"] and cols["stock"]:
            max_needed_col = max(v for v in cols.values() if v)
            for row in ws.iter_rows(min_row=header_row+1, min_col=1, max_col=max_needed_col, values_only=True):
                cod = row[cols["codigo"]-1] if cols["codigo"] else None
                if not _norm_text(cod): continue
                stock_raw = row[cols["stock"]-1] if cols["stock"] else None
                if not incluir_imsa(stock_raw, try_float(stock_raw)): 
                    continue
                precio = row[cols["precio"]-1] if cols["precio"] else None
                moneda = row[cols["moneda"]-1] if cols["moneda"] else None
                out.append({"ID": _norm_text(cod),
                            "Stock": convertir_stock_generico(stock_raw),
                            "Precio": try_float(precio), "Moneda": _norm_text(moneda) or None,
                            "StockRaw": stock_raw, "StockNum": try_float(stock_raw)})
        else:
            # Fallback si no detecta encabezados: suponemos stock en col H (8) como antes
            for row in ws.iter_rows(min_row=8, min_col=1, max_col=max(ws.max_column, 1), values_only=True):
                cod = row[0] if len(row) >= 1 else None
                if not _norm_text(cod): continue
                stock_raw = row[7] if len(row) >= 8 else None
                if not incluir_imsa(stock_raw, try_float(stock_raw)): 
                    continue
                out.append({"ID": _norm_text(cod),
                            "Stock": convertir_stock_generico(stock_raw),
                            "Precio": None, "Moneda": None,
                            "StockRaw": stock_raw, "StockNum": try_float(stock_raw)})
    finally:
        wb.close()
    return out

# ========= EXTRACTORES SOLO PARA DIFFS (mapeo fijo) =========
def extraer_tevelam(path: Path) -> List[Dict[str, Any]]:
    """Tevelam: desde fila 11; A=código (1), L=precio (12)."""
    out: List[Dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        max_row = ws.max_row or 1
        for r in range(11, max_row + 1):
            cod = ws.cell(row=r, column=1).value  # A
            if not _norm_text(cod): continue
            precio = ws.cell(row=r, column=12).value  # L
            out.append({"ID": _norm_text(cod), "Precio": try_float(precio), "Moneda": None})
    finally:
        wb.close()
    return out

def extraer_disco(path: Path) -> List[Dict[str, Any]]:
    """Disco: desde fila 8; A=código (1), M=precio (13)."""
    out: List[Dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        max_row = ws.max_row or 1
        for r in range(8, max_row + 1):
            cod = ws.cell(row=r, column=1).value  # A
            if not _norm_text(cod): continue
            precio = ws.cell(row=r, column=13).value  # M
            out.append({"ID": _norm_text(cod), "Precio": try_float(precio), "Moneda": None})
    finally:
        wb.close()
    return out

def extraer_imsa(path: Path) -> List[Dict[str, Any]]:
    """IMSA: desde fila 8; A=código (1), G=precio (7)."""
    out: List[Dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        max_row = ws.max_row or 1
        for r in range(8, max_row + 1):
            cod = ws.cell(row=r, column=1).value  # A
            if not _norm_text(cod): continue
            precio = ws.cell(row=r, column=7).value  # G
            out.append({"ID": _norm_text(cod), "Precio": try_float(precio), "Moneda": None})
    finally:
        wb.close()
    return out

# ========= SNAPSHOTS & REPORTES =========
def _snap_path(source_key: str) -> Path:
    return STATE_SNAP_DIR / f"{source_key}_snapshot.csv"

def guardar_snapshot(source_key: str, registros: List[Dict[str, Any]]) -> None:
    p = _snap_path(source_key)
    with p.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["ID","Precio","Moneda"])
        for r in registros:
            w.writerow([r.get("ID"), r.get("Precio"), r.get("Moneda")])
    try:
        (SNAP_DIR / p.name).write_bytes(p.read_bytes())
    except Exception:
        pass

def cargar_snapshot(source_key: str) -> Dict[str, Dict[str, Any]]:
    p = _snap_path(source_key)
    data: Dict[str, Dict[str, Any]] = {}
    if not p.exists():
        return data
    with p.open("r", newline="", encoding="utf-8") as f:
        rd = csv.DictReader(f)
        for row in rd:
            try:
                precio_val = float(row["Precio"]) if row["Precio"] not in (None, "", "None") else None
            except Exception:
                precio_val = None
            data[row["ID"]] = {"Precio": precio_val, "Moneda": row.get("Moneda") or None}
    return data

def calcular_diffs(prev_snap: Dict[str, Dict[str, Any]],
                   curr_regs: List[Dict[str, Any]]
                   ) -> Tuple[List[List[Any]], List[List[Any]], List[List[Any]], List[List[Any]]]:
    curr_map: Dict[str, Dict[str, Any]] = {r["ID"]: r for r in curr_regs if r.get("ID")}
    prev_ids = set(prev_snap.keys())
    curr_ids = set(curr_map.keys())

    nuevos_ids = sorted(curr_ids - prev_ids)
    elim_ids   = sorted(prev_ids - curr_ids)
    comunes    = sorted(curr_ids & prev_ids)

    precios_up: List[List[Any]] = []
    precios_dn: List[List[Any]] = []
    nuevos: List[List[Any]] = []
    eliminados: List[List[Any]] = []

    for _id in comunes:
        prev = prev_snap.get(_id, {})
        curr = curr_map.get(_id, {})
        p_old = prev.get("Precio")
        p_new = curr.get("Precio")
        mon   = curr.get("Moneda") or prev.get("Moneda")
        if p_old is None or p_new is None:
            continue
        if p_new != p_old:
            delta = p_new - p_old
            delta_pct = (delta / p_old * 100.0) if p_old != 0 else None
            row = [_id, mon, p_old, p_new, delta, delta_pct]
            if delta > 0: precios_up.append(row)
            else:         precios_dn.append(row)

    for _id in nuevos_ids:
        c = curr_map[_id]
        nuevos.append([_id, c.get("Moneda"), c.get("Precio")])

    for _id in elim_ids:
        p = prev_snap[_id]
        eliminados.append([_id, p.get("Moneda"), p.get("Precio")])

    return precios_up, precios_dn, nuevos, eliminados

def hay_cambios(precios_up, precios_dn, nuevos, eliminados) -> bool:
    return any([precios_up, precios_dn, nuevos, eliminados])

def crear_libro_cambios(source_key: str,
                        precios_up: List[List[Any]],
                        precios_dn: List[List[Any]],
                        nuevos: List[List[Any]],
                        eliminados: List[List[Any]]) -> Path:
    wb = Workbook(write_only=True)
    sh_res = wb.create_sheet("Resumen")
    sh_up  = wb.create_sheet("Precios ↑")
    sh_dn  = wb.create_sheet("Precios ↓")
    sh_new = wb.create_sheet("Nuevos modelos")
    sh_del = wb.create_sheet("Modelos eliminados")
    try:
        d = wb.worksheets[0]
        if d.title not in {"Resumen","Precios ↑","Precios ↓","Nuevos modelos","Modelos eliminados"}:
            wb.remove(d)
    except Exception:
        pass

    sh_up.append(["ID","Moneda","Precio anterior","Precio nuevo","Δ","Δ %"])
    sh_dn.append(["ID","Moneda","Precio anterior","Precio nuevo","Δ","Δ %"])
    sh_new.append(["ID","Moneda","Precio"])
    sh_del.append(["ID","Moneda","Precio"])

    for r in precios_up: sh_up.append(r)
    for r in precios_dn: sh_dn.append(r)
    for r in nuevos:     sh_new.append(r)
    for r in eliminados: sh_del.append(r)

    cnt_up = len(precios_up)
    cnt_dn = len(precios_dn)
    cnt_new = len(nuevos)
    cnt_del = len(eliminados)
    sum_up = round(sum((x[4] for x in precios_up)), 4) if cnt_up else 0
    sum_dn = round(sum((x[4] for x in precios_dn)), 4) if cnt_dn else 0

    sh_res.append(["Fuente", source_key])
    sh_res.append(["Generado", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    sh_res.append([])
    sh_res.append(["Métrica","Valor"])
    sh_res.append(["Precios ↑ (cantidad)", cnt_up])
    sh_res.append(["Precios ↓ (cantidad)", cnt_dn])
    sh_res.append(["Suma Δ ↑", sum_up])
    sh_res.append(["Suma Δ ↓", sum_dn])
    sh_res.append(["Nuevos modelos", cnt_new])
    sh_res.append(["Modelos eliminados", cnt_del])

    out = REPORTS_DIR / f"{source_key}_DIFF_{ts()}.xlsx"
    wb.save(out)
    log(f"🧾 Reporte generado: {out.name}")

    # copia pública
    try:
        (PUBLIC_REPORTS_DIR / out.name).write_bytes(out.read_bytes())
    except Exception as e:
        log(f"⚠️ No se pudo copiar reporte a public_reports: {e}")

    return out

# ========= SALIDA “Hoja 1” =========
def guardar_hoja1_xlsx(path_base: Path, registros: List[Dict[str, Any]], nombre_salida: Optional[str] = None) -> Path:
    wb_out = Workbook(write_only=True)
    h1 = wb_out.create_sheet("Hoja 1")
    try:
        d = wb_out.worksheets[0]
        if d.title != "Hoja 1":
            wb_out.remove(d)
    except Exception:
        pass
    h1.append(["ID","Stock","Precio","Moneda"])
    for r in registros:
        h1.append([r.get("ID"), r.get("Stock"), r.get("Precio"), r.get("Moneda")])
    out = path_base.with_name(path_base.stem + (nombre_salida or "_HOJA1") + ".xlsx")
    wb_out.save(out)
    log(f"✅ {path_base.stem} → {out.name}")

    # Copia pública (una por fuente)
    try:
        safe = f"{path_base.stem}_ULTIMA.xlsx"
        (PUBLIC_LISTAS_DIR / safe).write_bytes(out.read_bytes())
    except Exception as e:
        log(f"⚠️ No se pudo copiar Hoja 1 a public_listas: {e}")

    return out

# ========= Copia exacta DB + meta =========
def guardar_db_copia_exacta(source_key: str, path: Path, sha: Optional[str]) -> None:
    dst = PUBLIC_DB_DIR / f"{source_key}_DB.xlsx"
    try:
        dst.write_bytes(path.read_bytes())
    except Exception as e:
        log(f"⚠️ {source_key}: no se pudo copiar DB a public_db: {e}")
        return
    try:
        meta = {
            "source": source_key,
            "sha256": sha,
            "saved_at_utc": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "saved_at_ar": datetime.now(ZoneInfo("America/Argentina/Buenos_Aires")).strftime("%Y-%m-%d %H:%M:%S"),
            "filename": dst.name,
        }
        (PUBLIC_DB_DIR / f"{source_key}_DB.meta.json").write_text(
            json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8"
        )
    except Exception as e:
        log(f"⚠️ {source_key}: no se pudo escribir meta JSON: {e}")

# ========= SELENIUM (IMSA) =========
def _build_chrome() -> webdriver.Chrome:
    chrome_options = Options()
    chrome_options.add_argument("--headless=new")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
    chrome_options.add_argument("--window-size=1920,1080")
    prefs = {
        "download.default_directory": str(RUTA_DESCARGA),
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True,
    }
    chrome_options.add_experimental_option("prefs", prefs)
    service = Service(ChromeDriverManager().install())
    return webdriver.Chrome(service=service, options=chrome_options)

def _close_driver(driver: webdriver.Chrome):
    try: driver.quit()
    except Exception: pass

def _find_recent_listaimsa(max_age_sec: int = 180) -> Optional[Path]:
    now = time.time()
    pats = ("lista", "imsa")
    exts = (".xlsx", ".xls")
    candidatos = []
    for p in RUTA_DESCARGA.iterdir():
        if not p.is_file(): continue
        if p.suffix.lower() not in exts: continue
        name = p.name.lower()
        if not any(s in name for s in pats): continue
        try: mtime = p.stat().st_mtime
        except Exception: continue
        if now - mtime <= max_age_sec:
            if (RUTA_DESCARGA / (p.name + ".crdownload")).exists():
                continue
            candidatos.append(p)
    if not candidatos:
        return None
    return max(candidatos, key=lambda x: x.stat().st_mtime)

def descargar_imsa_web() -> Optional[Path]:
    driver = None
    try:
        driver = _build_chrome()
        log("🌐 Abriendo IMSA…")
        driver.get(IMSA_URL)

        try:
            iframes = WebDriverWait(driver, 20).until(EC.presence_of_all_elements_located((By.TAG_NAME, "iframe")))
            driver.switch_to.frame(iframes[0])
            WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, "pass"))).send_keys(IMSA_PASSWORD)
            btn = driver.find_element(By.XPATH, "//input[@type='submit' and @value='Login']")
            driver.execute_script("arguments[0].click();", btn)
            log("🔑 Login enviado.")
        except Exception as e:
            log(f"⚠️ No se pudo automatizar el login: {e}")
        finally:
            try: driver.switch_to.default_content()
            except Exception: pass

        log("⏳ Esperando 60s para descarga automática…")
        time.sleep(60)

        cand = _find_recent_listaimsa(max_age_sec=240)
        if cand:
            log(f"✅ IMSA detectado: {cand.name}")
            return cand

        log("🔎 Polling carpeta (hasta 60s)…")
        deadline = time.time() + 60
        while time.time() < deadline:
            cand = _find_recent_listaimsa(max_age_sec=240)
            if cand:
                log(f"✅ IMSA detectado: {cand.name}")
                return cand
            time.sleep(2)
        log("⚠️ No se detectó archivo IMSA.")
        return None
    except Exception as e:
        log(f"❌ Error en flujo IMSA: {e}")
        return None
    finally:
        if driver:
            _close_driver(driver)
            log("🧹 Selenium cerrado.")

# ========= MAIN =========
def run_fuente(source_key: str, path: Optional[Path],
               extractor_diffs, extractor_hoja1):
    if not path:
        return

    should, new_sha = decide_should_process(source_key, path)
    if not should:
        omitidos.append(source_key)
        return

    # 0) Copia exacta (DB) + meta pública
    guardar_db_copia_exacta(source_key, path, new_sha)

    # A) Stock V (Hoja 1)
    try:
        regs_h1 = extractor_hoja1(path)
        guardar_hoja1_xlsx(path, regs_h1)
    except Exception as e:
        log(f"⚠️ {source_key}: error generando Hoja 1: {e}")

    # B) Reporte de difs (evitar "primera vez")
    try:
        regs = extractor_diffs(path)
        prev = cargar_snapshot(source_key)
        first_time = (len(prev) == 0)
        if first_time:
            log(f"ℹ️ {source_key}: primer snapshot → no se genera reporte.")
        else:
            precios_up, precios_dn, nuevos, eliminados = calcular_diffs(prev, regs)
            if hay_cambios(precios_up, precios_dn, nuevos, eliminados):
                _ = crear_libro_cambios(source_key, precios_up, precios_dn, nuevos, eliminados)
            else:
                log(f"ℹ️ {source_key}: sin cambios de precio/modelos.")
        guardar_snapshot(source_key, regs)
    except Exception as e:
        log(f"⚠️ {source_key}: error calculando/generando diffs: {e}")

if __name__ == "__main__":
    log("INICIO: descarga + SHA completo + DB pública + Stock V + difs (estado persistente)")

    # 1) DESCARGAS
    try:
        log("Descargando Tevelam…")
        tevelam = download_simple(URL_TEVELAM, "Tevelam")
        log(f"Tevelam → {tevelam}")
    except Exception as e:
        log(f"ERROR Tevelam: {e}"); tevelam = None

    try:
        log("Descargando Disco Pro…")
        disco = download_simple(URL_DISCO_PRO, "Disco_Pro")
        log(f"Disco_Pro → {disco}")
    except Exception as e:
        log(f"ERROR Disco Pro: {e}"); disco = None

    try:
        log("Descargando Proveedor Extra…")
        extra = download_simple(URL_PROVEEDOR_EXTRA, "ARS_Tech")
        log(f"ARS_Tech → {extra}")
    except Exception as e:
        log(f"ERROR Proveedor Extra: {e}"); extra = None

    try:
        log("Lanzando IMSA…")
        imsa = descargar_imsa_web()
        log(f"IMSA detectado → {imsa}")
    except Exception as e:
        log(f"ERROR flujo IMSA: {e}"); imsa = None

    # 2) PROCESO (solo si SHA distinto)
    log("Procesando por fuente (según SHA)…")
    omitidos: List[str] = []

    if tevelam: run_fuente("Tevelam", tevelam, extraer_tevelam,        extraer_tevelam_hoja1)
    if disco:   run_fuente("Disco_Pro", disco, extraer_disco,          extraer_disco_hoja1)
    if extra:   run_fuente("ARS_Tech", extra, extraer_proveedor_extra, extraer_extra_hoja1)
    if imsa:    run_fuente("IMSA", imsa,    extraer_imsa,              extraer_imsa_hoja1)

    # 3) RESUMEN
    log("================ RESUMEN ================")
    if omitidos:
        log("Omitidos por hash igual (nada que hacer):")
        for n in omitidos:
            log(f"  • {n}")
    else:
        log("No hubo fuentes omitidas por hash igual.")

    log(f"FIN en: {RUTA_DESCARGA}")
